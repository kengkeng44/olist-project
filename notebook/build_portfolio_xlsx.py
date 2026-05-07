"""Generate output/portfolio.xlsx — 13-sheet industry-standard Excel portfolio.

Sheet layout (12 visible + 1 hidden):
  01_Cover, 02_TOC, 02_Data_Dictionary
  04_Executive_Summary
  03_KPI_Dashboard, 04_Revenue_Trend, 05_RFM_Segments, 06_Cohort_Heatmap
  07_ROI_Calculator (live formulas)
  08_Installments, 09_Logistics
  12_Methodology
  + hidden helper: _data_calc (central data hub; visible sheets pull values from here)

Industry conventions applied:
  - Excel Tables (ListObject) on every tabular block
  - Center Across Selection (centerContinuous) replaces merged cells for headers
  - Frozen panes on every data sheet
  - Conditional formatting (data bars / icon sets / color scale)
  - Named ranges for ROI inputs
  - Visible sheets reference _data_calc via formulas (XLOOKUP / direct ref)
  - No emoji; semantics conveyed via icon-set conditional formatting
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output"
DATA_DIR = ROOT / "data"
XLSX_PATH = OUTPUT_DIR / "portfolio.xlsx"

# ---------- Style palette ----------
NAVY = "1F3A5F"
TEAL = "2E8B8B"
ORANGE = "E07B39"
LIGHT_GRAY = "F2F2F2"
DARK_GRAY = "595959"
WHITE = "FFFFFF"
ACCENT_YELLOW = "F4D35E"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def font(*, size=11, bold=False, color="000000", italic=False) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def right() -> Alignment:
    return Alignment(horizontal="right", vertical="center", wrap_text=True)


def center_across() -> Alignment:
    """Center Across Selection — visual centering without merging (industry preferred)."""
    return Alignment(horizontal="centerContinuous", vertical="center", wrap_text=True)


def header_row(ws: Worksheet, row: int, cols: list[str], start_col: int = 2) -> None:
    for i, label in enumerate(cols):
        c = ws.cell(row=row, column=start_col + i, value=label)
        c.font = font(bold=True, color=WHITE)
        c.fill = fill(NAVY)
        c.alignment = center()
        c.border = BORDER


def title_block(ws: Worksheet, row: int, title: str, subtitle: str = "",
                start_col: int = 2, span: int = 5) -> None:
    """Main title at column B (industry margin convention), no merged cells."""
    end_col = start_col + span - 1
    c = ws.cell(row=row, column=start_col, value=title)
    c.font = font(size=18, bold=True, color=NAVY)
    c.alignment = center_across()
    for col in range(start_col + 1, end_col + 1):
        ws.cell(row=row, column=col).alignment = center_across()
    if subtitle:
        c2 = ws.cell(row=row + 1, column=start_col, value=subtitle)
        c2.font = font(size=11, italic=True, color=DARK_GRAY)
        c2.alignment = center_across()
        for col in range(start_col + 1, end_col + 1):
            ws.cell(row=row + 1, column=col).alignment = center_across()


def section_header(ws: Worksheet, row: int, text: str, start_col: int = 2) -> None:
    """Section header — always column B for vertical alignment with main title."""
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = font(size=13, bold=True, color=NAVY)


def add_table(ws: Worksheet, ref: str, name: str, style: str = "TableStyleMedium2") -> None:
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)


def add_named_range(wb: Workbook, name: str, sheet: str, ref: str) -> None:
    wb.defined_names.append(DefinedName(name=name, attr_text=f"'{sheet}'!{ref}"))


def set_col_widths(ws: Worksheet, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def read_csv(name: str) -> list[dict[str, str]]:
    with (OUTPUT_DIR / name).open(encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


# For each raw table, the column whose values we want to VARY across the
# 3 sample rows so reviewers see real diversity (not 3 identical-looking rows).
VARIETY_COL: dict[str, int] = {
    "olist_orders_dataset.csv": 2,            # order_status
    "olist_customers_dataset.csv": 4,         # customer_state
    "olist_order_items_dataset.csv": 1,       # order_item_id (1, 2, 3 → multi-item)
    "olist_order_payments_dataset.csv": 2,    # payment_type (credit_card / boleto / voucher)
    "olist_order_reviews_dataset.csv": 2,     # review_score (1 / 3 / 5)
}


def read_raw_sample(filename: str, n: int = 3) -> tuple[list[str], list[list[str]]]:
    """Read N "interesting" rows from a raw CSV — prefer rows where every
    field is non-empty AND vary the categorical column so reviewers see
    diversity instead of three identical-looking rows."""
    path = DATA_DIR / filename
    with path.open(encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = next(reader)
        all_rows = []
        for i, row in enumerate(reader):
            if i >= 2000:
                break
            all_rows.append(row)

    full_rows = [r for r in all_rows if all(v and v.strip() for v in r)]
    pool = full_rows if len(full_rows) >= n else (
        sorted(all_rows, key=lambda r: -sum(1 for v in r if v and v.strip()))
    )

    var_col = VARIETY_COL.get(filename)
    if var_col is not None and var_col < len(headers):
        seen: set[str] = set()
        picked: list[list[str]] = []
        for r in pool:
            v = r[var_col]
            if v not in seen:
                picked.append(r)
                seen.add(v)
                if len(picked) == n:
                    break
        # Pad if not enough unique values found
        if len(picked) < n:
            for r in pool:
                if r not in picked:
                    picked.append(r)
                    if len(picked) == n:
                        break
        return headers, picked[:n]
    return headers, pool[:n]


def count_csv_rows(filename: str) -> int:
    path = DATA_DIR / filename
    with path.open(encoding="utf-8-sig") as f:
        return sum(1 for _ in f) - 1  # exclude header


def aggregate_state_payment() -> tuple[list[str], list[str], dict[tuple[str, str], int]]:
    """Compute state × payment_type order count from raw CSVs.
    Returns (sorted_states, payment_types, dict[(state, ptype)] = count)."""
    from collections import defaultdict
    cust_state: dict[str, str] = {}
    with (DATA_DIR / "olist_customers_dataset.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            cust_state[row["customer_id"]] = row["customer_state"]

    order_state: dict[str, str] = {}
    with (DATA_DIR / "olist_orders_dataset.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            order_state[row["order_id"]] = cust_state.get(row["customer_id"], "")

    pivot: dict[tuple[str, str], int] = defaultdict(int)
    payment_types: set[str] = set()
    # Collapse duplicate payment lines: count each (state, ptype, order_id) at most once
    seen_per_combo: dict[tuple[str, str], set[str]] = defaultdict(set)
    with (DATA_DIR / "olist_order_payments_dataset.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            state = order_state.get(row["order_id"], "")
            ptype = row["payment_type"]
            if not state or not ptype or ptype == "not_defined":
                continue
            payment_types.add(ptype)
            combo = (state, ptype)
            if row["order_id"] not in seen_per_combo[combo]:
                pivot[combo] += 1
                seen_per_combo[combo].add(row["order_id"])

    # Sort states by total volume descending
    state_totals: dict[str, int] = defaultdict(int)
    for (s, _p), v in pivot.items():
        state_totals[s] += v
    states_sorted = sorted(state_totals.keys(), key=lambda s: state_totals[s], reverse=True)
    payment_sorted = sorted(payment_types, key=lambda p: -sum(pivot[(s, p)] for s in states_sorted))
    return states_sorted, payment_sorted, dict(pivot)


def read_category_translation() -> list[tuple[str, str]]:
    """Read PT→EN category translation. Returns list of (pt, en) tuples."""
    path = DATA_DIR / "product_category_name_translation.csv"
    out: list[tuple[str, str]] = []
    with path.open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            out.append((row["product_category_name"], row["product_category_name_english"]))
    return out


def aggregate_order_status() -> list[tuple[str, int]]:
    """Count orders by status from raw orders CSV. Returns sorted list."""
    from collections import Counter
    counter: Counter[str] = Counter()
    with (DATA_DIR / "olist_orders_dataset.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            counter[row["order_status"]] += 1
    # Roll minor statuses into 'other'
    keep = {"delivered", "shipped", "canceled", "unavailable"}
    out: dict[str, int] = {}
    for status, n in counter.items():
        if status in keep:
            out[status] = n
        else:
            out["other"] = out.get("other", 0) + n
    order = ["delivered", "shipped", "canceled", "unavailable", "other"]
    return [(s, out[s]) for s in order if s in out]


def aggregate_payment_mix() -> list[tuple[str, int, float]]:
    """Count payment_type and avg installments from raw payments CSV.
    Returns list of (payment_type, order_count, avg_installments) sorted by count desc."""
    from collections import defaultdict
    counts: dict[str, set] = defaultdict(set)
    inst_sum: dict[str, float] = defaultdict(float)
    inst_n: dict[str, int] = defaultdict(int)
    with (DATA_DIR / "olist_order_payments_dataset.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            ptype = row["payment_type"]
            if ptype == "not_defined":
                continue
            counts[ptype].add(row["order_id"])
            try:
                k = int(row["payment_installments"])
                inst_sum[ptype] += k
                inst_n[ptype] += 1
            except ValueError:
                pass
    out = []
    for p, ids in counts.items():
        n = len(ids)
        avg_inst = inst_sum[p] / inst_n[p] if inst_n[p] > 0 else 0.0
        out.append((p, n, round(avg_inst, 2)))
    out.sort(key=lambda x: -x[1])
    return out


# ============================================================================
# Hidden helper: _data_calc — central data hub
# ============================================================================
def build_data_calc(wb: Workbook) -> dict[str, int | str]:
    """Build the hidden _data_calc sheet. Returns row pointers so visible
    sheets can build cross-sheet formulas like =_data_calc!B5."""
    ws = wb.create_sheet("_data_calc")
    ws.sheet_state = "hidden"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 32, "B": 16, "C": 14, "D": 16, "E": 16, "F": 16,
                        "G": 14, "H": 14, "I": 12, "J": 10, "K": 10, "L": 10,
                        "M": 10, "N": 10, "O": 10, "P": 10})

    refs: dict[str, int | str] = {}

    # ---- Section 1: KPI Lookup ----
    ws.cell(row=1, column=1, value="KPI Lookup").font = font(bold=True, size=14, color=NAVY)
    header_row(ws, 2, ["Metric", "Value", "Format"], start_col=1)
    kpi_pairs = [
        ("Total orders", 99441, "#,##0"),
        ("Unique customers", 96096, "#,##0"),
        ("Sellers", 3095, "#,##0"),
        ("Products", 32951, "#,##0"),
        ("Product categories", 71, "#,##0"),
        ("States covered", 27, "0"),
        ("Cities covered", 4119, "#,##0"),
        ("Total GMV (R$)", 13591644.0, "#,##0"),
        ("AOV (R$)", 137.41, "0.00"),
        ("Cross-month repeat rate", 0.0181, "0.00%"),
        ("Champions ARPU (R$)", 274.0, "0"),
        ("At-Risk customers", 21975, "#,##0"),
        ("At-Risk ARPU (R$)", 213.0, "0"),
        ("Avg actual delivery (days)", 15.4, "0.0"),
        ("Avg ETA (days)", 24.0, "0.0"),
        ("Repeat customers (count)", 1693, "#,##0"),
    ]
    for i, (k, v, fmt) in enumerate(kpi_pairs):
        r = 3 + i
        ws.cell(row=r, column=1, value=k)
        c = ws.cell(row=r, column=2, value=v)
        c.number_format = fmt
        ws.cell(row=r, column=3, value=fmt).font = font(italic=True, size=9, color=DARK_GRAY)
    kpi_end = 2 + len(kpi_pairs)
    add_table(ws, f"A2:C{kpi_end}", "tbl_kpi_lookup")
    refs["kpi_start"] = 3
    refs["kpi_end"] = kpi_end

    # ---- Section 2: Yearly KPIs ----
    s2 = kpi_end + 3
    ws.cell(row=s2, column=1, value="Yearly KPIs").font = font(bold=True, size=14, color=NAVY)
    header_row(ws, s2 + 1, ["Year", "GMV (R$ M)", "Orders", "Avg Review", "Avg Delivery (days)"], start_col=1)
    yearly = read_csv("kpi_yearly.csv")
    for i, row in enumerate(yearly):
        r = s2 + 2 + i
        ws.cell(row=r, column=1, value=int(row["年份"]))
        ws.cell(row=r, column=2, value=float(row["總營收_M"]))
        ws.cell(row=r, column=3, value=int(row["總訂單數"]))
        ws.cell(row=r, column=4, value=float(row["平均評分"]))
        ws.cell(row=r, column=5, value=float(row["平均送達天數"]))
    y_end = s2 + 1 + len(yearly)
    add_table(ws, f"A{s2 + 1}:E{y_end}", "tbl_yearly")
    refs["yearly_start"] = s2 + 2
    refs["yearly_end"] = y_end

    # ---- Section 3: Revenue by month ----
    s3 = y_end + 3
    ws.cell(row=s3, column=1, value="Monthly Revenue").font = font(bold=True, size=14, color=NAVY)
    header_row(ws, s3 + 1, ["Month", "Revenue (R$)"], start_col=1)
    rev = read_csv("revenue.csv")
    for i, row in enumerate(rev):
        r = s3 + 2 + i
        ws.cell(row=r, column=1, value=row["月份"])
        ws.cell(row=r, column=2, value=float(row["總營收"]))
    r_end = s3 + 1 + len(rev)
    add_table(ws, f"A{s3 + 1}:B{r_end}", "tbl_revenue")
    refs["rev_start"] = s3 + 2
    refs["rev_end"] = r_end

    # ---- Section 4: RFM ----
    s4 = r_end + 3
    ws.cell(row=s4, column=1, value="RFM Segments").font = font(bold=True, size=14, color=NAVY)
    header_row(ws, s4 + 1, ["Segment", "Customers", "Customer %", "Revenue (R$)", "Revenue %", "ARPU (R$)", "Avg Recency (d)"], start_col=1)
    rfm = read_csv("rfm_segments.csv")
    seg_map = {
        "冠軍客戶": "Champions",
        "流失風險": "At Risk",
        "一般客戶": "Average",
        "忠誠客戶": "Loyal Low-spend",
        "已流失": "Lost",
        "潛力新客": "Potential New",
    }
    priority = ["冠軍客戶", "流失風險", "忠誠客戶", "一般客戶", "潛力新客", "已流失"]
    rfm_sorted = sorted(rfm, key=lambda r: priority.index(r["segment"]))
    for i, row in enumerate(rfm_sorted):
        r = s4 + 2 + i
        ws.cell(row=r, column=1, value=seg_map[row["segment"]])
        ws.cell(row=r, column=2, value=int(row["客戶數"]))
        ws.cell(row=r, column=3, value=float(row["客戶佔比_%"]) / 100)
        ws.cell(row=r, column=4, value=float(row["群組總營收"]))
        ws.cell(row=r, column=5, value=float(row["營收佔比_%"]) / 100)
        ws.cell(row=r, column=6, value=float(row["平均_M_元"]))
        ws.cell(row=r, column=7, value=float(row["平均_R_天"]))
    rfm_end = s4 + 1 + len(rfm_sorted)
    add_table(ws, f"A{s4 + 1}:G{rfm_end}", "tbl_rfm")
    refs["rfm_start"] = s4 + 2
    refs["rfm_end"] = rfm_end

    # ---- Section 5: Installments ----
    s5 = rfm_end + 3
    ws.cell(row=s5, column=1, value="Installments").font = font(bold=True, size=14, color=NAVY)
    header_row(ws, s5 + 1, ["Bucket", "Orders", "Avg Ticket (R$)", "Customers", "Repeat Rate"], start_col=1)
    inst = read_csv("installments.csv")
    bucket_en = {
        "1 (一次付清)": "1 (single)",
        "2-3 期": "2-3 installments",
        "4-6 期": "4-6 installments",
        "7-10 期": "7-10 installments",
        "11+ 期": "11+ installments",
    }
    for i, row in enumerate(inst):
        r = s5 + 2 + i
        ws.cell(row=r, column=1, value=bucket_en[row["bucket"]])
        ws.cell(row=r, column=2, value=int(row["orders"]))
        ws.cell(row=r, column=3, value=float(row["avg_ticket"]))
        ws.cell(row=r, column=4, value=int(row["customers"]))
        ws.cell(row=r, column=5, value=float(row["repeat_rate_pct"]) / 100)
    inst_end = s5 + 1 + len(inst)
    add_table(ws, f"A{s5 + 1}:E{inst_end}", "tbl_installments")
    refs["inst_start"] = s5 + 2
    refs["inst_end"] = inst_end

    # ---- Section 6: Logistics ----
    s6 = inst_end + 3
    ws.cell(row=s6, column=1, value="Logistics by State").font = font(bold=True, size=14, color=NAVY)
    header_row(ws, s6 + 1, ["State", "Actual days", "Estimated days", "ETA gap (days)"], start_col=1)
    log = read_csv("logistics.csv")
    for i, row in enumerate(log):
        r = s6 + 2 + i
        ws.cell(row=r, column=1, value=row["州"])
        actual = float(row["實際天數"])
        eta = float(row["預期天數"])
        ws.cell(row=r, column=2, value=actual)
        ws.cell(row=r, column=3, value=eta)
        ws.cell(row=r, column=4, value=eta - actual)
    log_end = s6 + 1 + len(log)
    add_table(ws, f"A{s6 + 1}:D{log_end}", "tbl_logistics")
    refs["log_start"] = s6 + 2
    refs["log_end"] = log_end

    # ---- Section 7: Cohort matrix ----
    s7 = log_end + 3
    ws.cell(row=s7, column=1, value="Cohort Retention Matrix").font = font(bold=True, size=14, color=NAVY)
    months = ["M0", "M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11", "M12"]
    header_row(ws, s7 + 1, ["Cohort"] + months, start_col=1)
    cohort = read_csv("cohort_retention.csv")
    for i, row in enumerate(cohort):
        r = s7 + 2 + i
        ws.cell(row=r, column=1, value=row["cohort_month"])
        for j, m in enumerate(months):
            v = row.get(m, "")
            if v:
                ws.cell(row=r, column=2 + j, value=float(v) / 100)
    cohort_end = s7 + 1 + len(cohort)
    last_col = get_column_letter(1 + len(months))
    add_table(ws, f"A{s7 + 1}:{last_col}{cohort_end}", "tbl_cohort")
    refs["cohort_start"] = s7 + 2
    refs["cohort_end"] = cohort_end

    # ---- Section 8: Order status (aggregated from raw orders CSV) ----
    s8 = cohort_end + 3
    ws.cell(row=s8, column=1, value="Order Status (raw aggregation)").font = font(bold=True, size=14, color=NAVY)
    header_row(ws, s8 + 1, ["Status", "Orders", "Share"], start_col=1)
    status_rows = aggregate_order_status()
    total_orders = sum(n for _, n in status_rows)
    refs["status_start"] = s8 + 2
    for i, (s_name, n) in enumerate(status_rows):
        r = s8 + 2 + i
        ws.cell(row=r, column=1, value=s_name)
        c = ws.cell(row=r, column=2, value=n); c.number_format = "#,##0"
        # Share computed via formula (visible function: simple division)
        sc = ws.cell(row=r, column=3, value=f"=B{r}/SUM($B${s8+2}:$B${s8+1+len(status_rows)})")
        sc.number_format = "0.0%"
    refs["status_end"] = s8 + 1 + len(status_rows)
    add_table(ws, f"A{s8+1}:C{refs['status_end']}", "tbl_status_raw")

    # ---- Section 9: Payment mix (aggregated from raw payments CSV) ----
    s9 = refs["status_end"] + 3
    ws.cell(row=s9, column=1, value="Payment Mix (raw aggregation)").font = font(bold=True, size=14, color=NAVY)
    header_row(ws, s9 + 1, ["Payment type", "Orders", "Share", "Avg installments"], start_col=1)
    pay_rows = aggregate_payment_mix()
    refs["payment_start"] = s9 + 2
    for i, (p, n, avg) in enumerate(pay_rows):
        r = s9 + 2 + i
        ws.cell(row=r, column=1, value=p)
        c = ws.cell(row=r, column=2, value=n); c.number_format = "#,##0"
        sc = ws.cell(row=r, column=3, value=f"=B{r}/SUM($B${s9+2}:$B${s9+1+len(pay_rows)})")
        sc.number_format = "0.0%"
        ic = ws.cell(row=r, column=4, value=avg); ic.number_format = "0.0"
    refs["payment_end"] = s9 + 1 + len(pay_rows)
    add_table(ws, f"A{s9+1}:D{refs['payment_end']}", "tbl_payment_raw")

    # ---- Section 10: Derived headlines (formula-computed, used by Cover) ----
    s10 = refs["payment_end"] + 3
    ws.cell(row=s10, column=1, value="Derived Headlines (formulas)").font = font(bold=True, size=14, color=NAVY)
    header_row(ws, s10 + 1, ["Metric", "Formula result"], start_col=1)
    # Look up specific values via VLOOKUP / INDEX / direct ref so the formulas are visible
    rfm_atrisk_count_formula = (
        f"=VLOOKUP(\"At Risk\",$A${refs['rfm_start']}:$G${refs['rfm_end']},2,FALSE)"
    )
    rfm_atrisk_arpu_formula = (
        f"=VLOOKUP(\"At Risk\",$A${refs['rfm_start']}:$G${refs['rfm_end']},6,FALSE)"
    )
    inst_aov_710_formula = (
        f"=VLOOKUP(\"7-10 installments\",$A${refs['inst_start']}:$E${refs['inst_end']},3,FALSE)"
    )
    inst_aov_1_formula = (
        f"=VLOOKUP(\"1 (single)\",$A${refs['inst_start']}:$E${refs['inst_end']},3,FALSE)"
    )

    headline_metrics = [
        ("At-Risk count (from RFM)", rfm_atrisk_count_formula, "#,##0"),
        ("At-Risk ARPU (from RFM)", rfm_atrisk_arpu_formula, "0"),
        ("7-10 inst AOV (from Installments)", inst_aov_710_formula, "0"),
        ("1-inst AOV (from Installments)", inst_aov_1_formula, "0"),
        # Derived: aggressive ROI revenue = at_risk × 20% × ARPU × 50%
        ("Aggressive ROI revenue (R$)",
         f"=B{s10+2}*0.2*B{s10+3}*0.5", "#,##0"),
        # Derived: AOV ratio 7-10 / 1
        ("AOV ratio (7-10 / 1)",
         f"=B{s10+4}/B{s10+5}", "0.00"),
        # Repeat rate (already in KPI lookup) — pull via XLOOKUP
        ("Cross-month repeat rate",
         f'=_xlfn.XLOOKUP("Cross-month repeat rate",$A$3:$A${refs["kpi_end"]},$B$3:$B${refs["kpi_end"]})',
         "0.00%"),
    ]
    for i, (label, fml, fmt) in enumerate(headline_metrics):
        r = s10 + 2 + i
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value=fml); c.number_format = fmt
    refs["headline_start"] = s10 + 2
    refs["headline_end"] = s10 + 1 + len(headline_metrics)
    add_table(ws, f"A{s10+1}:B{refs['headline_end']}", "tbl_headlines")

    # Specific cell pointers for Cover formulas
    refs["cell_aggressive_roi"] = f"_data_calc!$B${s10+6}"
    refs["cell_aov_ratio"] = f"_data_calc!$B${s10+7}"
    refs["cell_repeat_rate"] = f"_data_calc!$B${s10+8}"

    return refs


def kpi_lookup_formula(refs: dict, metric_name: str) -> str:
    """XLOOKUP formula fetching a KPI value from _data_calc by metric name.
    Uses the `_xlfn.` prefix that openpyxl requires for modern Excel functions."""
    range_metric = f"_data_calc!$A$3:$A${refs['kpi_end']}"
    range_value = f"_data_calc!$B$3:$B${refs['kpi_end']}"
    return f'=_xlfn.XLOOKUP("{metric_name}",{range_metric},{range_value})'


# ============================================================================
# 01_Cover
# ============================================================================
def build_cover(wb: Workbook, refs: dict) -> None:
    ws = wb.create_sheet("01_Cover")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 2, "B": 28, "C": 28, "D": 28, "E": 28, "F": 2})

    ws.row_dimensions[2].height = 40
    title = ws.cell(row=2, column=2, value="Olist Brazilian E-Commerce")
    title.font = font(size=26, bold=True, color=NAVY)
    title.alignment = center_across()
    for col in range(3, 6):
        ws.cell(row=2, column=col).alignment = center_across()

    ws.row_dimensions[3].height = 26
    sub = ws.cell(row=3, column=2,
                  value="End-to-End Customer Analytics — SQL · RFM · Cohort · ROI")
    sub.font = font(size=14, italic=True, color=DARK_GRAY)
    sub.alignment = center_across()
    for col in range(3, 6):
        ws.cell(row=3, column=col).alignment = center_across()

    # Author / role / dataset block
    pairs = [
        ("Author", "Jen-Ho Cheng"),
        ("Role", "Data / Product Analyst — E-commerce"),
        ("Dataset", "Olist (Brazil) · 99,441 orders · 2016–2018"),
    ]
    for i, (k, v) in enumerate(pairs):
        ws.cell(row=5 + i, column=2, value=k).font = font(bold=True, color=NAVY)
        ws.cell(row=5 + i, column=3, value=v).font = font()

    # Headline metrics — formulas pulling from _data_calc Derived Headlines
    section_header(ws, 9, "THREE HEADLINE FINDINGS (formula-driven)")
    # Big-number cells use TEXT() to format the formula result inline as the
    # display string the reviewer sees. Each big number is computed live.
    big_formulas = [
        f'="R$ "&TEXT({refs["cell_aggressive_roi"]}/1000,"0")&"K"',
        f'=TEXT({refs["cell_repeat_rate"]},"0.00%")',
        f'=TEXT({refs["cell_aov_ratio"]},"0.00")&"×"',
    ]
    mid_labels = [
        "Win-back revenue opportunity",
        "Cross-month repeat rate",
        "ARPU lift from 7-10 installments",
    ]
    small_labels = [
        "21,975 at-risk customers · 9.4× ROI aggressive scenario",
        "Platform-level retention failure (vs. 5-15% mature benchmark)",
        "R$ 334 vs R$ 96 single payment — Brazil's hidden CRM",
    ]
    for i in range(3):
        col = 2 + i
        ws.row_dimensions[11].height = 38
        ws.row_dimensions[12].height = 22
        ws.row_dimensions[13].height = 36
        c = ws.cell(row=11, column=col, value=big_formulas[i])
        c.font = font(size=24, bold=True, color=WHITE)
        c.fill = fill([NAVY, TEAL, ORANGE][i])
        c.alignment = center()
        c.border = BORDER
        c2 = ws.cell(row=12, column=col, value=mid_labels[i])
        c2.font = font(size=11, bold=True, color=WHITE)
        c2.fill = fill([NAVY, TEAL, ORANGE][i])
        c2.alignment = center()
        c2.border = BORDER
        c3 = ws.cell(row=13, column=col, value=small_labels[i])
        c3.font = font(size=9)
        c3.fill = fill(LIGHT_GRAY)
        c3.alignment = center()
        c3.border = BORDER

    # ---- TOC (merged from old 02_TOC) ----
    section_header(ws, 16, "SHEETS IN THIS WORKBOOK")
    header_row(ws, 17, ["#", "Sheet", "What it shows"])
    toc_entries = [
        ("02", "02_Data_Dictionary", "9-table schema + 3-row samples"),
        ("03", "03_KPI_Dashboard", "Scale / Yearly KPIs / Status / Payment — XLOOKUP"),
        ("04", "04_Revenue_Trend", "Monthly revenue + line chart"),
        ("05", "05_RFM_Segments", "RFM segments (NTILE) + Pareto + icon set"),
        ("06", "06_Cohort_Heatmap", "13-month retention matrix (color scale)"),
        ("07", "07_ROI_Calculator", "Live formulas + Named Ranges"),
        ("08", "08_Installments", "AOV + repeat rate by installment bucket"),
        ("09", "09_Logistics", "State ETA gap with 3-color scale"),
        ("10", "10_Pivot_Analysis", "Real PivotTable: state x payment_type"),
    ]
    for i, (num, sheet, desc) in enumerate(toc_entries):
        r = 18 + i
        ws.cell(row=r, column=2, value=num).alignment = center()
        c = ws.cell(row=r, column=3, value=sheet)
        c.font = font(bold=True, color="0563C1")
        c.hyperlink = f"#'{sheet}'!A1"
        c.alignment = left()
        ws.cell(row=r, column=4, value=desc).alignment = left()
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).border = BORDER

    # ---- Live links ----
    links_row = 18 + len(toc_entries) + 2
    section_header(ws, links_row, "LIVE LINKS")
    links = [
        ("Streamlit App", "https://olist-jenho.streamlit.app/",
         "Interactive 5-page dashboard"),
        ("HTML Dashboard", "https://kengkeng44.github.io/olist-project/",
         "Static GitHub Pages"),
        ("Tableau Public",
         "https://public.tableau.com/app/profile/jenho.cheng/viz/2_17739060990590/1?publish=yes",
         "Tableau version"),
        ("GitHub Repo", "https://github.com/kengkeng44/olist-project",
         "Full source + SQL + notebook"),
        ("Pitch Deck", "https://github.com/kengkeng44/olist-project/blob/master/slides/portfolio.pdf",
         "13-slide PM deck"),
    ]
    header_row(ws, links_row + 1, ["Resource", "URL", "Description"])
    for i, (label, url, desc) in enumerate(links):
        r = links_row + 2 + i
        ws.cell(row=r, column=2, value=label).font = font(bold=True)
        c = ws.cell(row=r, column=3, value=url)
        c.hyperlink = url
        c.font = font(color="0563C1")
        c.alignment = left()
        ws.cell(row=r, column=4, value=desc).alignment = left()
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).border = BORDER

    # ---- Tech stack chips ----
    stack_row = links_row + 2 + len(links) + 2
    section_header(ws, stack_row, "TECH STACK")
    stack = ["Python", "SQLite", "SQL (NTILE)", "pandas", "matplotlib", "Streamlit", "Tableau"]
    for i, t in enumerate(stack):
        c = ws.cell(row=stack_row + 1 + i // 4, column=2 + i % 4, value=t)
        c.font = font(bold=True, color=NAVY)
        c.fill = fill(ACCENT_YELLOW)
        c.alignment = center()
        c.border = BORDER


# ============================================================================
# 02_Data_Dictionary
# ============================================================================
def build_data_dictionary(wb: Workbook) -> None:
    ws = wb.create_sheet("02_Data_Dictionary")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 2, "B": 26, "C": 14, "D": 44, "E": 2})

    title_block(ws, 2, "Data Dictionary",
                "9 tables · 1.55M rows. 3-row samples below.",
                span=4)

    # Schema overview — one-line descriptions only
    section_header(ws, 5, "9 source tables")
    header_row(ws, 6, ["Table", "Rows", "What's in it"])
    schema = [
        ("olist_orders_dataset", "olist_orders_dataset.csv",
         "Order header + 8 timestamps (purchase → delivery)."),
        ("olist_customers_dataset", "olist_customers_dataset.csv",
         "customer_id (per order) + customer_unique_id + ZIP."),
        ("olist_order_items_dataset", "olist_order_items_dataset.csv",
         "Line items: 1 row per (order, product, seller) with price + freight."),
        ("olist_order_payments_dataset", "olist_order_payments_dataset.csv",
         "Payment type (boleto / credit_card / voucher / debit_card) + installments."),
        ("olist_order_reviews_dataset", "olist_order_reviews_dataset.csv",
         "1-5 star reviews + comment text."),
        ("olist_products_dataset", "olist_products_dataset.csv",
         "Product attributes (category, weight, dimensions)."),
        ("olist_sellers_dataset", "olist_sellers_dataset.csv",
         "Seller ZIP + state."),
        ("olist_geolocation_dataset", "olist_geolocation_dataset.csv",
         "ZIP-level lat/lng (state-level aggregations only)."),
        ("product_category_name_translation", "product_category_name_translation.csv",
         "PT->EN mapping for the 71 product categories."),
    ]
    for i, (name, fname, desc) in enumerate(schema):
        r = 7 + i
        rows = count_csv_rows(fname)
        ws.cell(row=r, column=2, value=name).font = Font(name="Consolas", size=10, bold=True)
        ws.cell(row=r, column=3, value=rows).number_format = "#,##0"
        ws.cell(row=r, column=4, value=desc).alignment = left()
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).border = BORDER
    schema_end = 6 + len(schema)
    add_table(ws, f"B6:D{schema_end}", "tbl_schema")

    # Sample blocks — 3 rows each, the 5 most-joined tables
    sample_files = [
        ("orders", "olist_orders_dataset.csv"),
        ("customers", "olist_customers_dataset.csv"),
        ("items", "olist_order_items_dataset.csv"),
        ("payments", "olist_order_payments_dataset.csv"),
        ("reviews", "olist_order_reviews_dataset.csv"),
    ]
    cur_row = schema_end + 3
    for short, filename in sample_files:
        section_header(ws, cur_row, f"{filename} — first 3 rows")
        cur_row += 1
        headers, rows = read_raw_sample(filename, n=3)
        for i, _ in enumerate(headers):
            col = 2 + i
            ws.column_dimensions[get_column_letter(col)].width = max(
                14, ws.column_dimensions[get_column_letter(col)].width or 0
            )
        header_row(ws, cur_row, headers)
        for ri, row in enumerate(rows):
            r = cur_row + 1 + ri
            for ci, val in enumerate(row):
                # Try numeric coercion so Excel stops flagging "number stored as text"
                typed: int | float | str = val
                if val and val.lstrip("-").replace(".", "", 1).isdigit():
                    typed = float(val) if "." in val else int(val)
                cell = ws.cell(row=r, column=2 + ci, value=typed)
                cell.font = Font(name="Consolas", size=9)
                cell.border = BORDER
                cell.alignment = left()
        sample_end = cur_row + len(rows)
        last_col = get_column_letter(1 + len(headers))
        add_table(ws, f"B{cur_row}:{last_col}{sample_end}", f"tbl_sample_{short}")
        cur_row = sample_end + 2

    ws.freeze_panes = "B7"


# ============================================================================
# 03_KPI_Dashboard — formulas reference _data_calc
# ============================================================================
def build_kpis(wb: Workbook, refs: dict) -> None:
    ws = wb.create_sheet("03_KPI_Dashboard")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 2, "B": 32, "C": 18, "D": 18, "E": 18, "F": 18, "G": 2})

    title_block(ws, 2, "Data Overview & KPIs", span=5)

    # ---- Scale & Coverage (formula-driven) ----
    section_header(ws, 5, "Scale & Coverage")
    header_row(ws, 6, ["Metric", "Value"])
    overview_metrics = [
        ("Total orders", "#,##0"),
        ("Unique customers", "#,##0"),
        ("Sellers", "#,##0"),
        ("Products", "#,##0"),
        ("Product categories", "#,##0"),
        ("States covered", "0"),
        ("Cities covered", "#,##0"),
        ("Total GMV (R$)", "#,##0"),
        ("AOV (R$)", "0.00"),
    ]
    for i, (k, fmt) in enumerate(overview_metrics):
        r = 7 + i
        ws.cell(row=r, column=2, value=k).font = font(bold=True)
        ws.cell(row=r, column=2).fill = fill(LIGHT_GRAY)
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).alignment = left()
        c = ws.cell(row=r, column=3, value=kpi_lookup_formula(refs, k))
        c.number_format = fmt
        c.alignment = right()
        c.border = BORDER

    overview_end = 6 + len(overview_metrics)

    # ---- Yearly KPIs — XLOOKUP by year (visible function, year-order independent) ----
    yk_row = overview_end + 3
    section_header(ws, yk_row, "Year-over-Year KPIs")
    header_row(ws, yk_row + 1, ["Year", "GMV (R$ M)", "Orders", "Avg Review", "Avg Delivery (days)"])
    yearly = read_csv("kpi_yearly.csv")
    src_year_keys = f"_data_calc!$A${refs['yearly_start']}:$A${refs['yearly_end']}"
    yk_cols = ["B", "C", "D", "E"]   # GMV, Orders, Review, Delivery
    yk_fmts = ["0.00", "#,##0", "0.00", "0.00"]
    years = [int(row["年份"]) for row in yearly]
    for i, yr in enumerate(years):
        r = yk_row + 2 + i
        # Year (literal — XLOOKUP key)
        ws.cell(row=r, column=2, value=yr).border = BORDER
        ws.cell(row=r, column=2).alignment = left()
        # 4 metric columns via XLOOKUP keyed by year
        for ci, (src_col, fmt) in enumerate(zip(yk_cols, yk_fmts)):
            src_range = f"_data_calc!${src_col}${refs['yearly_start']}:${src_col}${refs['yearly_end']}"
            cell = ws.cell(row=r, column=3 + ci,
                           value=f'=_xlfn.XLOOKUP(B{r},{src_year_keys},{src_range},0)')
            cell.number_format = fmt
            cell.border = BORDER
            cell.alignment = center()

    # ---- Order Status — pulled via XLOOKUP + SUMIFS from _data_calc ----
    st_row = yk_row + 2 + len(yearly) + 2
    section_header(ws, st_row, "Order Status Distribution (from raw)")
    header_row(ws, st_row + 1, ["Status", "Orders", "Share"])
    statuses_known = ["delivered", "shipped", "canceled", "unavailable", "other"]
    status_a_range = f"_data_calc!$A${refs['status_start']}:$A${refs['status_end']}"
    status_b_range = f"_data_calc!$B${refs['status_start']}:$B${refs['status_end']}"
    status_c_range = f"_data_calc!$C${refs['status_start']}:$C${refs['status_end']}"
    for i, s in enumerate(statuses_known):
        r = st_row + 2 + i
        ws.cell(row=r, column=2, value=s).border = BORDER
        # XLOOKUP the order count
        c = ws.cell(row=r, column=3,
                    value=f'=_xlfn.XLOOKUP("{s}",{status_a_range},{status_b_range},0)')
        c.number_format = "#,##0"; c.border = BORDER
        # XLOOKUP the share
        c2 = ws.cell(row=r, column=4,
                     value=f'=_xlfn.XLOOKUP("{s}",{status_a_range},{status_c_range},0)')
        c2.number_format = "0.0%"; c2.border = BORDER
    st_end = st_row + 1 + len(statuses_known)
    add_table(ws, f"B{st_row + 1}:D{st_end}", "tbl_status")

    # ---- Payment Mix — pulled via XLOOKUP from _data_calc ----
    pm_row = st_end + 3
    section_header(ws, pm_row, "Payment Mix (from raw)")
    header_row(ws, pm_row + 1, ["Payment type", "Orders", "Share", "Avg installments"])
    payments_known = ["credit_card", "boleto", "voucher", "debit_card"]
    pay_a_range = f"_data_calc!$A${refs['payment_start']}:$A${refs['payment_end']}"
    pay_b_range = f"_data_calc!$B${refs['payment_start']}:$B${refs['payment_end']}"
    pay_c_range = f"_data_calc!$C${refs['payment_start']}:$C${refs['payment_end']}"
    pay_d_range = f"_data_calc!$D${refs['payment_start']}:$D${refs['payment_end']}"
    for i, p in enumerate(payments_known):
        r = pm_row + 2 + i
        ws.cell(row=r, column=2, value=p).border = BORDER
        c = ws.cell(row=r, column=3,
                    value=f'=_xlfn.XLOOKUP("{p}",{pay_a_range},{pay_b_range},0)')
        c.number_format = "#,##0"; c.border = BORDER
        c2 = ws.cell(row=r, column=4,
                     value=f'=_xlfn.XLOOKUP("{p}",{pay_a_range},{pay_c_range},0)')
        c2.number_format = "0.0%"; c2.border = BORDER
        c3 = ws.cell(row=r, column=5,
                     value=f'=_xlfn.XLOOKUP("{p}",{pay_a_range},{pay_d_range},0)')
        c3.number_format = "0.0"; c3.border = BORDER
    pm_end = pm_row + 1 + len(payments_known)
    add_table(ws, f"B{pm_row + 1}:E{pm_end}", "tbl_payment")

    ws.freeze_panes = "B5"


# ============================================================================
# 04_Revenue_Trend — formulas + Excel Table + chart
# ============================================================================
def build_revenue(wb: Workbook, refs: dict) -> None:
    ws = wb.create_sheet("04_Revenue_Trend")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 2, "B": 14, "C": 18, "D": 2})

    title_block(ws, 2, "Monthly Revenue Trend (2017–2018)", span=2)

    rev = read_csv("revenue.csv")
    header_row(ws, 5, ["Month", "Revenue (R$)"])
    src_months = f"_data_calc!$A${refs['rev_start']}:$A${refs['rev_end']}"
    src_rev = f"_data_calc!$B${refs['rev_start']}:$B${refs['rev_end']}"
    months = [row["月份"] for row in rev]
    for i, month in enumerate(months):
        r = 6 + i
        ws.cell(row=r, column=2, value=month).border = BORDER
        cell = ws.cell(row=r, column=3, value=f'=_xlfn.XLOOKUP(B{r},{src_months},{src_rev},0)')
        cell.number_format = "#,##0"
        cell.border = BORDER
    end_row = 5 + len(months)
    add_table(ws, f"B5:C{end_row}", "tbl_revenue_view")

    # Line chart
    chart = LineChart()
    chart.title = "Monthly Revenue (R$) — peak Nov 2017 (Black Friday + Christmas)"
    chart.y_axis.title = "Revenue (R$)"
    chart.x_axis.title = "Month"
    chart.height = 10
    chart.width = 22
    data = Reference(ws, min_col=3, min_row=5, max_row=end_row, max_col=3)
    cats = Reference(ws, min_col=2, min_row=6, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.style = 12
    ws.add_chart(chart, "E5")

    ws.freeze_panes = "B6"


# ============================================================================
# 05_RFM_Segments — formulas, no emoji, icon set conditional formatting
# ============================================================================
def build_rfm(wb: Workbook, refs: dict) -> None:
    ws = wb.create_sheet("05_RFM_Segments")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 2, "B": 22, "C": 14, "D": 14, "E": 14, "F": 14, "G": 18, "H": 14, "I": 32, "J": 2})

    title_block(ws, 2, "RFM Customer Segmentation",
                "Champions + At-Risk = 40% customers, 67% revenue.",
                span=8)

    persona_map = {
        "Champions": "Big-ticket buyer, ordered last month — protect, upsell",
        "At Risk": "High-value customer dormant ~13 months — top win-back priority",
        "Average": "Mid-frequency, mid-value. Cross-sell test bed",
        "Loyal Low-spend": "Recent + frequent but low ticket. Bundle / category expansion",
        "Lost": "Long inactive + low value. Lowest priority",
        "Potential New": "Just acquired, low ticket. Onboarding nurture",
    }

    header_row(ws, 5, ["Segment", "Customers", "Customer %", "Revenue (R$)",
                       "Revenue %", "ARPU (R$)", "Avg Recency (d)", "Persona / Action"])

    # Read data to get count + segment names in priority order
    rfm = read_csv("rfm_segments.csv")
    seg_map = {
        "冠軍客戶": "Champions", "流失風險": "At Risk", "一般客戶": "Average",
        "忠誠客戶": "Loyal Low-spend", "已流失": "Lost", "潛力新客": "Potential New",
    }
    priority = ["冠軍客戶", "流失風險", "忠誠客戶", "一般客戶", "潛力新客", "已流失"]
    rfm_sorted = sorted(rfm, key=lambda r: priority.index(r["segment"]))

    src_seg = f"_data_calc!$A${refs['rfm_start']}:$A${refs['rfm_end']}"
    rfm_cols = ["B", "C", "D", "E", "F", "G"]   # Customers, %, Rev, %, ARPU, Recency
    rfm_fmts = ["#,##0", "0.0%", "#,##0", "0.0%", "0", "0"]
    for i, row in enumerate(rfm_sorted):
        r = 6 + i
        en = seg_map[row["segment"]]
        # Segment name — XLOOKUP key
        ws.cell(row=r, column=2, value=en).font = font(bold=True)
        ws.cell(row=r, column=2).border = BORDER
        # 6 numeric columns each via XLOOKUP keyed by segment name
        for ci, (src_col, fmt) in enumerate(zip(rfm_cols, rfm_fmts)):
            src_range = f"_data_calc!${src_col}${refs['rfm_start']}:${src_col}${refs['rfm_end']}"
            cell = ws.cell(row=r, column=3 + ci,
                           value=f'=_xlfn.XLOOKUP(B{r},{src_seg},{src_range},0)')
            cell.number_format = fmt
            cell.border = BORDER
        # Persona text — literal
        p = ws.cell(row=r, column=9, value=persona_map[en])
        p.alignment = left()
        p.border = BORDER

    end_row = 5 + len(rfm_sorted)
    add_table(ws, f"B5:I{end_row}", "tbl_rfm_view")

    # Conditional formatting:
    # - Data bar on Customer %
    # - Data bar on Revenue %
    # - 3-symbol icon set on ARPU (high/mid/low ARPU)
    ws.conditional_formatting.add(
        f"D6:D{end_row}",
        DataBarRule(start_type="min", end_type="max", color=TEAL,
                    showValue=True),
    )
    ws.conditional_formatting.add(
        f"F6:F{end_row}",
        DataBarRule(start_type="min", end_type="max", color=ORANGE,
                    showValue=True),
    )
    ws.conditional_formatting.add(
        f"G6:G{end_row}",
        IconSetRule("3TrafficLights1", "percent", [0, 33, 67], showValue=True),
    )

    # Bar chart: customer % vs revenue %
    chart = BarChart()
    chart.type = "bar"
    chart.style = 12
    chart.title = "Customer share vs Revenue share — Pareto: Champions+At-Risk = 39.7% / 66.9%"
    chart.y_axis.title = "Segment"
    chart.x_axis.title = "Share"
    chart.height = 10
    chart.width = 22
    data1 = Reference(ws, min_col=4, min_row=5, max_row=end_row, max_col=4)
    data2 = Reference(ws, min_col=6, min_row=5, max_row=end_row, max_col=6)
    cats = Reference(ws, min_col=2, min_row=6, max_row=end_row)
    chart.add_data(data1, titles_from_data=True)
    chart.add_data(data2, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"B{end_row + 3}")

    ws.freeze_panes = "B6"


# ============================================================================
# 06_Cohort_Heatmap — frozen panes + color scale
# ============================================================================
def build_cohort(wb: Workbook, refs: dict) -> None:
    ws = wb.create_sheet("06_Cohort_Heatmap")
    ws.sheet_view.showGridLines = False

    title_block(ws, 2, "Cohort Retention Heatmap",
                "Cell = % of cohort active in month N (benchmark M1: 5-15%).",
                span=14)

    months = ["M0", "M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11", "M12"]
    header_row(ws, 5, ["Cohort"] + months)
    set_col_widths(ws, {"A": 2, "B": 12} | {get_column_letter(i): 8 for i in range(3, 16)})

    cohort = read_csv("cohort_retention.csv")
    cohort_keys = [row["cohort_month"] for row in cohort]
    # INDEX(MATCH(...), MATCH(...)) — picks the cell at (row=cohort_month, col=month_label)
    # in the _data_calc cohort matrix. Visible 2D lookup pattern.
    matrix_range = (f"_data_calc!$B${refs['cohort_start']}:"
                    f"${get_column_letter(1 + len(months))}${refs['cohort_end']}")
    cohort_key_range = f"_data_calc!$A${refs['cohort_start']}:$A${refs['cohort_end']}"
    month_header_range = (f"_data_calc!$B${refs['cohort_start']-1}:"
                          f"${get_column_letter(1 + len(months))}${refs['cohort_start']-1}")
    for i, ck in enumerate(cohort_keys):
        r = 6 + i
        # Cohort month (literal — MATCH key)
        ws.cell(row=r, column=2, value=ck).font = font(bold=True)
        ws.cell(row=r, column=2).fill = fill(LIGHT_GRAY)
        ws.cell(row=r, column=2).border = BORDER
        for j in range(len(months)):
            cell = ws.cell(row=r, column=3 + j,
                           value=(f"=INDEX({matrix_range},"
                                  f"MATCH($B{r},{cohort_key_range},0),"
                                  f"MATCH({get_column_letter(3+j)}$5,{month_header_range},0))"))
            cell.number_format = "0.0%"
            cell.alignment = center()
            cell.border = BORDER
            cell.font = font(size=9)

    end_row = 5 + len(cohort_keys)
    add_table(ws, f"B5:O{end_row}", "tbl_cohort_view")

    # Color scale on M1+ (skip M0 which is always 100%)
    rule = ColorScaleRule(
        start_type="num", start_value=0, start_color="FFFFFF",
        mid_type="num", mid_value=0.05, mid_color=ACCENT_YELLOW,
        end_type="num", end_value=0.10, end_color=ORANGE,
    )
    ws.conditional_formatting.add(f"D6:O{end_row}", rule)

    ws.freeze_panes = "C6"


# ============================================================================
# 07_ROI_Calculator — Named Ranges + live formulas
# ============================================================================
def build_roi(wb: Workbook, refs: dict) -> None:
    ws = wb.create_sheet("07_ROI_Calculator")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 2, "B": 32, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 2})

    title_block(ws, 2, "At-Risk Win-back ROI Calculator",
                "Edit yellow cells — scenarios recompute live.",
                span=6)

    # XLOOKUP-derived defaults pulled from RFM (data-driven) — user can still
    # type over these to model different segments.
    rfm_seg_range = f"_data_calc!$A${refs['rfm_start']}:$A${refs['rfm_end']}"
    rfm_count_range = f"_data_calc!$B${refs['rfm_start']}:$B${refs['rfm_end']}"
    rfm_arpu_range = f"_data_calc!$F${refs['rfm_start']}:$F${refs['rfm_end']}"
    at_risk_count_formula = f'=_xlfn.XLOOKUP("At Risk",{rfm_seg_range},{rfm_count_range},0)'
    at_risk_arpu_formula = f'=_xlfn.XLOOKUP("At Risk",{rfm_seg_range},{rfm_arpu_range},0)'

    # ---- Inputs ----
    section_header(ws, 5, "INPUTS — edit me")
    header_row(ws, 6, ["Parameter", "Value", "Format", "Notes"])
    inputs = [
        ("At-risk customers (segment size)", at_risk_count_formula, "#,##0",
         "XLOOKUP from RFM segment", "At_Risk_Count"),
        ("Avg ARPU per customer (R$)", at_risk_arpu_formula, "0",
         "XLOOKUP from RFM segment", "ARPU"),
        ("Repeat-spend rate (% of ARPU on win-back order)", 0.50, "0%",
         "Modeling assumption (50% of historical avg)", "Repeat_Spend"),
        ("Total CRM cost (R$)", 50000, "#,##0",
         "Budget assumption (email + discount voucher)", "CRM_Cost"),
    ]
    for i, (label, val, fmt, note, _name) in enumerate(inputs):
        r = 7 + i
        ws.cell(row=r, column=2, value=label).font = font(bold=True)
        c = ws.cell(row=r, column=3, value=val)
        c.fill = fill(ACCENT_YELLOW)
        c.font = font(bold=True, size=12)
        c.alignment = center()
        c.number_format = fmt
        medium_orange = Side(style="medium", color=ORANGE)
        c.border = Border(left=medium_orange, right=medium_orange,
                          top=medium_orange, bottom=medium_orange)
        ws.cell(row=r, column=4, value=fmt).font = font(italic=True, size=9, color=DARK_GRAY)
        ws.cell(row=r, column=5, value=note).font = font(italic=True, size=9, color=DARK_GRAY)
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=4).border = BORDER
        ws.cell(row=r, column=5).border = BORDER

    # Named Ranges for inputs (point at the value cells C7..C10)
    add_named_range(wb, "At_Risk_Count", "07_ROI_Calculator", "$C$7")
    add_named_range(wb, "ARPU", "07_ROI_Calculator", "$C$8")
    add_named_range(wb, "Repeat_Spend", "07_ROI_Calculator", "$C$9")
    add_named_range(wb, "CRM_Cost", "07_ROI_Calculator", "$C$10")

    # ---- Scenarios ----
    section_header(ws, 13, "SCENARIO OUTPUTS — recomputes live")
    header_row(ws, 14, ["Metric", "Conservative (5%)", "Optimistic (10%)",
                        "Aggressive (20%)", "Custom"])

    # Scenario columns aligned with header columns: C=Conservative, D=Optimistic,
    # E=Aggressive, F=Custom (yellow = editable). Off-by-one bug fixed here.
    scenarios = [("C", 0.05), ("D", 0.10), ("E", 0.20), ("F", 0.10)]
    custom_col = "F"

    # Recall rate row
    ws.cell(row=15, column=2, value="Recall rate (%)").font = font(bold=True)
    ws.cell(row=15, column=2).border = BORDER
    for col_letter, default in scenarios:
        col_idx = ord(col_letter) - ord("A") + 1
        c = ws.cell(row=15, column=col_idx, value=default)
        c.fill = fill(ACCENT_YELLOW if col_letter == custom_col else LIGHT_GRAY)
        c.number_format = "0%"
        c.alignment = center()
        c.font = font(bold=True)
        c.border = BORDER

    # Recalled customers — uses Named Range At_Risk_Count
    ws.cell(row=16, column=2, value="Recalled customers").font = font(bold=True)
    ws.cell(row=16, column=2).border = BORDER
    for col_letter, _ in scenarios:
        col_idx = ord(col_letter) - ord("A") + 1
        cell = ws.cell(row=16, column=col_idx, value=f"=At_Risk_Count*{col_letter}15")
        cell.number_format = "#,##0"
        cell.alignment = center()
        cell.border = BORDER

    # Incremental revenue — uses Named Ranges
    ws.cell(row=17, column=2, value="Incremental revenue (R$)").font = font(bold=True)
    ws.cell(row=17, column=2).border = BORDER
    for col_letter, _ in scenarios:
        col_idx = ord(col_letter) - ord("A") + 1
        cell = ws.cell(row=17, column=col_idx,
                       value=f"=At_Risk_Count*{col_letter}15*ARPU*Repeat_Spend")
        cell.number_format = "#,##0"
        cell.alignment = center()
        cell.border = BORDER

    # CRM cost
    ws.cell(row=18, column=2, value="CRM cost (R$)").font = font(bold=True)
    ws.cell(row=18, column=2).border = BORDER
    for col_letter, _ in scenarios:
        col_idx = ord(col_letter) - ord("A") + 1
        cell = ws.cell(row=18, column=col_idx, value="=CRM_Cost")
        cell.number_format = "#,##0"
        cell.alignment = center()
        cell.border = BORDER

    # Net = revenue - cost
    ws.cell(row=19, column=2, value="Net contribution (R$)").font = font(bold=True)
    ws.cell(row=19, column=2).border = BORDER
    for col_letter, _ in scenarios:
        col_idx = ord(col_letter) - ord("A") + 1
        cell = ws.cell(row=19, column=col_idx, value=f"={col_letter}17-{col_letter}18")
        cell.number_format = "#,##0"
        cell.alignment = center()
        cell.border = BORDER
        cell.font = font(bold=True)

    # ROI
    ws.cell(row=20, column=2, value="ROI (×)").font = font(bold=True, color=WHITE)
    ws.cell(row=20, column=2).fill = fill(NAVY)
    ws.cell(row=20, column=2).border = BORDER
    for col_letter, _ in scenarios:
        col_idx = ord(col_letter) - ord("A") + 1
        cell = ws.cell(row=20, column=col_idx, value=f"={col_letter}17/{col_letter}18")
        cell.number_format = "0.0\"×\""
        cell.alignment = center()
        cell.font = font(bold=True, size=14, color=WHITE)
        cell.fill = fill(TEAL)
        cell.border = BORDER

    # Formula explanation
    section_header(ws, 22, "FORMULA")
    ws.cell(row=23, column=2,
            value="Incremental revenue  =  At_Risk_Count × Recall rate × ARPU × Repeat_Spend").font = font(italic=True)
    ws.cell(row=24, column=2,
            value="ROI                  =  Incremental revenue ÷ CRM_Cost").font = font(italic=True)



# ============================================================================
# 08_Installments — formulas, icon set
# ============================================================================
def build_installments(wb: Workbook, refs: dict) -> None:
    ws = wb.create_sheet("08_Installments")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 2, "B": 22, "C": 14, "D": 18, "E": 14, "F": 18, "G": 2})

    title_block(ws, 2, "Installments — AOV & Repeat Rate by Bucket",
                "73.9% of orders use credit_card, avg 3.5 installments.",
                span=5)

    inst = read_csv("installments.csv")
    header_row(ws, 5, ["Installment bucket", "Orders", "Avg ticket (R$)", "Customers", "Repeat rate"])
    # Source ranges in _data_calc — use XLOOKUP keyed by bucket name so the
    # function is visible and order-independent.
    src_buckets = f"_data_calc!$A${refs['inst_start']}:$A${refs['inst_end']}"
    src_orders = f"_data_calc!$B${refs['inst_start']}:$B${refs['inst_end']}"
    src_aov = f"_data_calc!$C${refs['inst_start']}:$C${refs['inst_end']}"
    src_cust = f"_data_calc!$D${refs['inst_start']}:$D${refs['inst_end']}"
    src_repeat = f"_data_calc!$E${refs['inst_start']}:$E${refs['inst_end']}"
    bucket_labels = ["1 (single)", "2-3 installments", "4-6 installments",
                     "7-10 installments", "11+ installments"]
    for i, label in enumerate(bucket_labels):
        r = 6 + i
        # Bucket name (literal — Excel needs the key for XLOOKUP)
        ws.cell(row=r, column=2, value=label).border = BORDER
        # XLOOKUP for each metric — bucket as key
        ws.cell(row=r, column=3, value=f'=_xlfn.XLOOKUP(B{r},{src_buckets},{src_orders},0)').border = BORDER
        ws.cell(row=r, column=3).number_format = "#,##0"
        ws.cell(row=r, column=4, value=f'=_xlfn.XLOOKUP(B{r},{src_buckets},{src_aov},0)').border = BORDER
        ws.cell(row=r, column=4).number_format = "0"
        ws.cell(row=r, column=5, value=f'=_xlfn.XLOOKUP(B{r},{src_buckets},{src_cust},0)').border = BORDER
        ws.cell(row=r, column=5).number_format = "#,##0"
        ws.cell(row=r, column=6, value=f'=_xlfn.XLOOKUP(B{r},{src_buckets},{src_repeat},0)').border = BORDER
        ws.cell(row=r, column=6).number_format = "0.00%"

    end_row = 5 + len(bucket_labels)
    add_table(ws, f"B5:F{end_row}", "tbl_inst_view")

    # Conditional formatting: data bar on Avg ticket, icon set on Repeat rate
    ws.conditional_formatting.add(
        f"D6:D{end_row}",
        DataBarRule(start_type="min", end_type="max", color=TEAL, showValue=True),
    )
    ws.conditional_formatting.add(
        f"F6:F{end_row}",
        IconSetRule("3TrafficLights1", "percent", [0, 33, 67], showValue=True),
    )

    # AOV bar chart
    bar = BarChart()
    bar.type = "col"
    bar.style = 12
    bar.title = "Avg ticket (R$) by installment bucket — 7-10 installments = 3.48× single-payment"
    bar.y_axis.title = "Avg ticket (R$)"
    bar.x_axis.title = "Installments"
    bar.height = 9
    bar.width = 18
    data = Reference(ws, min_col=4, min_row=5, max_row=end_row, max_col=4)
    cats = Reference(ws, min_col=2, min_row=6, max_row=end_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, f"B{end_row + 3}")

    # Repeat rate bar chart
    bar2 = BarChart()
    bar2.type = "col"
    bar2.style = 13
    bar2.title = "Repeat rate by installment bucket — 7+ installment customers retain 65% better"
    bar2.y_axis.title = "Repeat rate"
    bar2.x_axis.title = "Installments"
    bar2.height = 9
    bar2.width = 18
    data2 = Reference(ws, min_col=6, min_row=5, max_row=end_row, max_col=6)
    bar2.add_data(data2, titles_from_data=True)
    bar2.set_categories(cats)
    ws.add_chart(bar2, f"B{end_row + 22}")

    ws.freeze_panes = "B6"


# ============================================================================
# 09_Logistics — formulas + 3-color scale
# ============================================================================
def build_logistics(wb: Workbook, refs: dict) -> None:
    ws = wb.create_sheet("09_Logistics")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 2, "B": 12, "C": 18, "D": 18, "E": 18, "F": 2})

    title_block(ws, 2, "Logistics — Estimated vs Actual Delivery (by State)",
                "National avg: 15.4 actual / 24 ETA days (36% under-promise).",
                span=4)

    log = read_csv("logistics.csv")
    header_row(ws, 5, ["State", "Actual days", "Estimated days", "ETA gap (days)"])
    # XLOOKUP by state code — order-independent, visible function
    src_states = f"_data_calc!$A${refs['log_start']}:$A${refs['log_end']}"
    src_actual = f"_data_calc!$B${refs['log_start']}:$B${refs['log_end']}"
    src_eta = f"_data_calc!$C${refs['log_start']}:$C${refs['log_end']}"
    src_gap = f"_data_calc!$D${refs['log_start']}:$D${refs['log_end']}"
    # Pull state codes from raw log csv (they're the keys for XLOOKUP)
    state_codes = [row["州"] for row in log]
    for i, state in enumerate(state_codes):
        r = 6 + i
        ws.cell(row=r, column=2, value=state).font = font(bold=True)
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=3, value=f'=_xlfn.XLOOKUP(B{r},{src_states},{src_actual},0)').border = BORDER
        ws.cell(row=r, column=3).number_format = "0.0"
        ws.cell(row=r, column=4, value=f'=_xlfn.XLOOKUP(B{r},{src_states},{src_eta},0)').border = BORDER
        ws.cell(row=r, column=4).number_format = "0.0"
        ws.cell(row=r, column=5, value=f'=_xlfn.XLOOKUP(B{r},{src_states},{src_gap},0)').border = BORDER
        ws.cell(row=r, column=5).number_format = "0.0"

    end_row = 5 + len(state_codes)
    add_table(ws, f"B5:E{end_row}", "tbl_logistics_view")

    # 3-color scale on ETA gap
    rule = ColorScaleRule(
        start_type="min", start_color="63BE7B",  # green = small gap (well-calibrated)
        mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
        end_type="max", end_color="F8696B",  # red = huge gap (over-promised buffer)
    )
    ws.conditional_formatting.add(f"E6:E{end_row}", rule)

    # Data bars on actual + estimated
    ws.conditional_formatting.add(
        f"C6:C{end_row}",
        DataBarRule(start_type="min", end_type="max", color=TEAL, showValue=True),
    )
    ws.conditional_formatting.add(
        f"D6:D{end_row}",
        DataBarRule(start_type="min", end_type="max", color=ORANGE, showValue=True),
    )

    # Bar chart
    chart = BarChart()
    chart.type = "bar"
    chart.style = 12
    chart.title = "Actual vs Estimated delivery days — SP best (8.8d), RN worst (19.3d)"
    chart.height = 14
    chart.width = 22
    data = Reference(ws, min_col=3, min_row=5, max_row=end_row, max_col=4)
    cats = Reference(ws, min_col=2, min_row=6, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "G5")

    ws.freeze_panes = "B6"


# ============================================================================
# 10_Pivot_Analysis — real PivotTable (built post-save via win32com)
# ============================================================================
def build_pivot_analysis(wb: Workbook) -> None:
    ws = wb.create_sheet("10_Pivot_Analysis")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {
        "A": 2, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14,
        "H": 4, "I": 14, "J": 14, "K": 12, "L": 2,
    })

    title_block(ws, 2, "PivotTable — Orders by State × Payment Type", span=10)

    # Long-form source (right side) — feeds the PivotCache
    states, ptypes, pivot_data = aggregate_state_payment()
    ws.cell(row=4, column=9, value="Source: tbl_payments_long").font = font(bold=True, color=NAVY)
    header_row(ws, 5, ["State", "Payment type", "Orders"], start_col=9)
    long_rows: list[tuple[str, str, int]] = []
    for s in states:
        for p in ptypes:
            v = pivot_data.get((s, p), 0)
            if v > 0:
                long_rows.append((s, p, v))
    for i, (s, p, v) in enumerate(long_rows):
        r = 6 + i
        ws.cell(row=r, column=9, value=s).border = BORDER
        ws.cell(row=r, column=10, value=p).border = BORDER
        c = ws.cell(row=r, column=11, value=v); c.number_format = "#,##0"; c.border = BORDER
    long_end = 5 + len(long_rows)
    add_table(ws, f"I5:K{long_end}", "tbl_payments_long")

    # PivotTable will be inserted at B5 by post-build (add_real_pivottable)
    ws.cell(row=4, column=2, value="PivotTable (built by Excel COM post-save)").font = font(bold=True, color=NAVY)

    ws.freeze_panes = "B5"


# ============================================================================
# Post-processor: add a real PivotTable via Excel COM (win32com)
# ============================================================================
def add_real_pivottable(file_path: Path) -> None:
    """Open the saved xlsx in Excel, build a PivotTable on 10_Pivot_Analysis
    using the tbl_payments_long source, save and quit. Windows + Excel only."""
    import win32com.client as w  # local import: only needed at build time

    xlRowField, xlColumnField, xlDataField = 1, 2, 4
    xlSum = -4157
    xlDatabase = 1

    excel = w.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(file_path.resolve()))
        ws = wb.Sheets("10_Pivot_Analysis")

        # Source = the structured table tbl_payments_long
        cache = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData="tbl_payments_long")
        # Destination cell — top-left of the visible pivot area
        dest = ws.Range("B5")
        pt = cache.CreatePivotTable(TableDestination=dest, TableName="pt_state_payment")

        pt.PivotFields("State").Orientation = xlRowField
        pt.PivotFields("Payment type").Orientation = xlColumnField
        data_field = pt.PivotFields("Orders")
        data_field.Orientation = xlDataField
        # Rename + sum
        try:
            pt.PivotFields("Sum of Orders").Function = xlSum
            pt.PivotFields("Sum of Orders").NumberFormat = "#,##0"
        except Exception:
            pass

        # Optional: style
        try:
            pt.TableStyle2 = "PivotStyleMedium9"
        except Exception:
            pass

        wb.Save()
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()
        del excel


# ============================================================================
# Build
# ============================================================================
def main() -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # Hidden helper first so other builders can reference its row layout
    refs = build_data_calc(wb)

    # Visible sheets in display order
    build_cover(wb, refs)
    build_data_dictionary(wb)
    build_kpis(wb, refs)
    build_revenue(wb, refs)
    build_rfm(wb, refs)
    build_cohort(wb, refs)
    build_roi(wb, refs)
    build_installments(wb, refs)
    build_logistics(wb, refs)
    build_pivot_analysis(wb)

    sheet_order = [
        "01_Cover", "02_Data_Dictionary",
        "03_KPI_Dashboard", "04_Revenue_Trend", "05_RFM_Segments", "06_Cohort_Heatmap",
        "07_ROI_Calculator", "08_Installments", "09_Logistics", "10_Pivot_Analysis",
        "_data_calc",
    ]
    wb._sheets = [wb[name] for name in sheet_order]
    wb.active = 0
    wb.save(XLSX_PATH)
    print(f"Wrote base xlsx: {XLSX_PATH} ({XLSX_PATH.stat().st_size / 1024:.1f} KB)")

    # Post-process: open in Excel via COM and add real PivotTable
    try:
        add_real_pivottable(XLSX_PATH)
        print(f"Added real PivotTable. Final size: {XLSX_PATH.stat().st_size / 1024:.1f} KB")
    except Exception as e:
        print(f"WARN: PivotTable post-build failed ({type(e).__name__}: {e})."
              f" Base xlsx is still valid; you can add the pivot manually.")


if __name__ == "__main__":
    main()
